from openpyxl import load_workbook

wb = load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
cell = sheet["a1"]
print(cell.row)
print(cell.column)
print(cell.coordinate)