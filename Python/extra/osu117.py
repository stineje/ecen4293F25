from openpyxl import load_workbook

wb = load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
cell = sheet["a1"]
print(sheet.max_row)
print(sheet.max_column)
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column):
        cell = sheet.cell(row, column)
        print(cell.value)