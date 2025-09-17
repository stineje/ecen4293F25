from openpyxl import load_workbook

wb = load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
cells = sheet[1:4]
sheet.append([1004, "Holliday", 8.95])
sheet.insert_rows(5)
wb.save("transaction2.xlsx")