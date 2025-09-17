from openpyxl import load_workbook

wb = load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)
sheet.append([1005, "Robertson", 9.95])
wb.save("transaction2.xlsx")

